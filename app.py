import os
import csv
import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import webbrowser
from datetime import datetime
import time

def search_in_excel(directory, search_terms):
    results = []

    # Iterate through all Excel files in the directory
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(root, file)
                try:
                    workbook = load_workbook(file_path, data_only=True)

                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]

                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            for col_idx, cell in enumerate(row, start=1):
                                for term in search_terms:
                                    if term == str(cell):
                                        cell_location = f"{sheet.cell(row=row_idx, column=col_idx).coordinate}"
                                        results.append((term, file_path, sheet_name, cell_location))
                                        break
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")

    return results

def browse_directory():
    directory = filedialog.askdirectory()
    if directory:
        dir_entry.delete(0, ctk.END)
        dir_entry.insert(0, directory)

def update_counters(input_count, result_count, search_time):
    input_counter_label.configure(text=f"Input Lines: {input_count}")
    result_counter_label.configure(text=f"Results Found: {result_count}")
    time_taken_label.configure(text=f"Time Taken: {search_time:.2f} sec")

def update_input_counter(event=None):
    search_terms = search_entry.get("1.0", ctk.END).strip().splitlines()
    input_count = len(search_terms)
    input_counter_label.configure(text=f"Input Lines: {input_count}")

def search():
    directory = dir_entry.get()
    search_terms = search_entry.get("1.0", ctk.END).strip().splitlines()

    if not directory:
        messagebox.showerror("Error", "Please select a directory.")
        return

    if not search_terms:
        messagebox.showerror("Error", "Please enter at least one search term.")
        return

    clear_results()
    update_counters(len(search_terms), 0, 0)
    
    # Start timer for search duration
    start_time = time.time()
    results = search_in_excel(directory, search_terms)
    end_time = time.time()

    # Calculate time taken to search
    search_time = end_time - start_time
    update_counters(len(search_terms), len(results), search_time)

    if results:
        for result in results:
            tree.insert("", ctk.END, values=result)
    else:
        messagebox.showinfo("No Results", "No matches found.")

def clear_results():
    for item in tree.get_children():
        tree.delete(item)
    update_counters(0, 0, 0)

def save_results():
    if not tree.get_children():
        messagebox.showwarning("Warning", "No results to save.")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = f"results_{timestamp}.csv"

    with open(file_path, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["Term", "File Path", "Sheet Name", "Cell Location"])
        for item in tree.get_children():
            writer.writerow(tree.item(item, "values"))

    messagebox.showinfo("Success", f"Results saved to {file_path}.")

def open_excel_file(event):
    selected_item = tree.selection()
    if selected_item:
        file_path = tree.item(selected_item, "values")[1]
        webbrowser.open(file_path)

def toggle_mode():
    if appearance_switch.get():
        ctk.set_appearance_mode("light")
    else:
        ctk.set_appearance_mode("dark")

# Initialize main window
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

root = ctk.CTk()
root.title("Excel Search Tool")
root.geometry("800x600")

# Directory selection
frame_dir = ctk.CTkFrame(root)
frame_dir.pack(pady=10, padx=10, fill="x")

dir_label = ctk.CTkLabel(frame_dir, text="Select Directory:", font=("Arial", 10, "bold"))
dir_label.pack(side="left", padx=5)

default_dir = os.getcwd()
dir_entry = ctk.CTkEntry(frame_dir, width=400)
dir_entry.insert(0, default_dir)
dir_entry.pack(side="left", padx=5, fill="x", expand=True)

browse_button = ctk.CTkButton(frame_dir, text="Browse", command=browse_directory)
browse_button.pack(side="left", padx=5)

# Search terms input
frame_search = ctk.CTkFrame(root)
frame_search.pack(pady=10, padx=10, fill="both", expand=True)

search_label = ctk.CTkLabel(frame_search, text="Enter Search Terms (one per line):", font=("Arial", 10, "bold"))
search_label.pack(anchor="w", padx=5)

search_entry = ctk.CTkTextbox(frame_search, height=100)
search_entry.pack(padx=5, pady=5, fill="both", expand=True)
search_entry.bind("<KeyRelease>", update_input_counter)

# Search, clear, and save buttons
frame_buttons = ctk.CTkFrame(root)
frame_buttons.pack(pady=10)

search_button = ctk.CTkButton(frame_buttons, text="Search", command=search)
search_button.pack(side="left", padx=5)

clear_button = ctk.CTkButton(frame_buttons, text="Clear Results", command=clear_results)
clear_button.pack(side="left", padx=5)

save_button = ctk.CTkButton(frame_buttons, text="Save Results", command=save_results)
save_button.pack(side="left", padx=5)

# Counters
frame_counters = ctk.CTkFrame(root)
frame_counters.pack(pady=5)

input_counter_label = ctk.CTkLabel(frame_counters, text="Input Lines: 0", font=("Arial", 10, "bold"))
input_counter_label.pack(side="left", padx=10)

result_counter_label = ctk.CTkLabel(frame_counters, text="Results Found: 0", font=("Arial", 10, "bold"))
result_counter_label.pack(side="left", padx=10)

time_taken_label = ctk.CTkLabel(frame_counters, text="Time Taken: 0.00 sec", font=("Arial", 10, "bold"))
time_taken_label.pack(side="left", padx=10)

# Results display
frame_results = ctk.CTkFrame(root)
frame_results.pack(pady=10, padx=10, fill="both", expand=True)

columns = ("Term", "File Path", "Sheet Name", "Cell Location")
tree_scrollbar = ttk.Scrollbar(frame_results, orient="vertical")
tree_scrollbar.pack(side="right", fill="y")

tree = ttk.Treeview(frame_results, columns=columns, show="headings", yscrollcommand=tree_scrollbar.set)
tree_scrollbar.config(command=tree.yview)

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=200)

tree.pack(padx=5, pady=5, fill="both", expand=True)
tree.bind("<Double-1>", open_excel_file)

# Footer
footer_frame = ctk.CTkFrame(root)
footer_frame.pack(side="bottom", fill="x", pady=10)

footer_label = ctk.CTkLabel(footer_frame, text="Made with ❤ by Aessa", font=("Arial", 12, "italic"))
footer_label.pack(side="left", padx=10)

# Appearance switch
appearance_switch = ctk.CTkSwitch(footer_frame, text="Dark/Light Mode", command=toggle_mode)
appearance_switch.pack(side="right", padx=10)

# Start the application
root.mainloop()
